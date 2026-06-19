"""
Carga y procesamiento de datos financieros desde archivos Excel de Merkacol.
Los archivos están hardcodeados para la demo; en producción se parametrizan por empresa.
"""
from pathlib import Path
import openpyxl
import pandas as pd
from functools import lru_cache

MERKACOL_DIR = Path("/home/crispy/Projects/merkacol-project")
FUENTE_FILE  = MERKACOL_DIR / "LIBRO MAYOR RESUMIDO.xlsm"
PYG_FILE     = MERKACOL_DIR / "PYG 2025 X TRIMESTRES COMPLETO.xlsm"

COMP_MONTH_COLS = {1:5,2:8,3:11,4:14,5:17,6:20,7:29,8:32,9:35,10:38,11:41,12:44}
MERK_MONTH_COLS = {1:5,2:8,3:11,4:19,5:22,6:25,7:38,8:41,9:44,10:52,11:55,12:58}
MERK_QUARTER_COLS = {1:14,2:28,3:47,4:61}

MESES = {1:"Ene",2:"Feb",3:"Mar",4:"Abr",5:"May",6:"Jun",
         7:"Jul",8:"Ago",9:"Sep",10:"Oct",11:"Nov",12:"Dic"}

PYG_STRUCTURE = [
    (0, "INGRESOS OPERACIONALES", []),
    (1, "VENTA MENSUAL", []),
    (2, "FRUVER", []),
    (3, "Vta Agrícolas (Excluidos)",    ["4135220000"]),
    (2, "CÁRNICOS", []),
    (3, "Vta Pecuarios (Exentos)",      ["4135230000"]),
    (3, "Vta Pecuarios (IVA 16%)",      ["4135240000"]),
    (3, "Vta Pecuarios (IVA 19%)",      ["4135250000"]),
    (2, "PGC", []),
    (3, "Vta PGC (Excluidos)",          ["4135350500"]),
    (3, "Vta PGC (Exentos)",            ["4135351000"]),
    (3, "Vta PGC (IVA 16%)",            ["4135351500"]),
    (3, "Vta PGC (IVA 5%)",             ["4135352000"]),
    (3, "Vta PGC (IVA 19%)",            ["4135352500"]),
    (2, "DOMICILIOS", []),
    (3, "Servicio Domicilio",           ["4135800000"]),
    (2, "ARRENDAMIENTOS", []),
    (3, "Arrend. Bienes Inmuebles",     ["4155050000"]),
    (1, "DEVOLUCIÓN EN VENTAS", []),
    (3, "Devoluciones (DB)",            ["4175100000"]),
    (0, "COSTOS", []),
    (2, "FRUVER", []),
    (3, "Inv. Inicial Fruver",          ["6201010000"]),
    (3, "Compras",                      ["6210050500"]),
    (3, "Traslados Entrada",            ["6210051000"]),
    (3, "Traslados Salida",             ["6210051500"]),
    (3, "Inv. Final Fruver",            ["6230010000"]),
    (2, "CÁRNICOS", []),
    (3, "Inv. Inicial Cárnicos",        ["6240050000"]),
    (3, "Compras Cárnicos (Exentos)",   ["6240100000"]),
    (3, "Compras Cárnicos (Excluidos)", ["6240090000"]),
    (3, "Compras Cárnicos (IVA 19%)",   ["6240110000"]),
    (3, "Inv. Final Cárnicos",          ["6240150000"]),
    (3, "Costo Vta Cárnicos",           ["6241050000"]),
    (3, "Ajustes Inventario",           ["6241100000"]),
    (3, "Bajas y Averías",              ["6241150000"]),
    (2, "PGC", []),
    (3, "Inv. Inicial PGC",             ["6242050000"]),
    (3, "Compras PGC (Exentos)",        ["6242200000"]),
    (3, "Compras PGC (Excluidos)",      ["6242210000"]),
    (3, "Compras PGC (5%)",             ["6242220000"]),
    (3, "Compras PGC (19%)",            ["6242240000"]),
    (3, "Inv. Final PGC",               ["6242100000"]),
    (3, "Costo Vta PGC",                ["6250050000"]),
    (3, "Ajustes Inventario PGC",       ["6250100000"]),
    (3, "Bajas y Averías PGC",          ["6250150000"]),
    (0, "GASTOS OPERACIONALES", []),
    (1, "DE PERSONAL", []),
    (2, "SUELDOS", []),
    (3, "Nómina",                       ["5205060100"]),
    (3, "Horas Extras",                 ["5205150000"]),
    (3, "Aux. Transporte",              ["5205270000"]),
    (3, "Aux. No Salarial",             ["5205450000"]),
    (3, "Bonificaciones",               ["5205480000"]),
    (2, "APORTES SOCIALES", []),
    (3, "Cesantías",                    ["5205300000"]),
    (3, "Int. Cesantías",               ["5205330000"]),
    (3, "Prima Servicios",              ["5205360000"]),
    (3, "Vacaciones",                   ["5205390000"]),
    (3, "Dotación",                     ["5205510000"]),
    (3, "Capacitación",                 ["5205630000"]),
    (1, "FINANCIEROS", []),
    (3, "Gastos Bancarios",             ["5305050100"]),
    (3, "Impuesto 4x1000",              ["5305050200"]),
    (3, "Intereses",                    ["5305200000"]),
    (1, "DIRECTOS DE OPERACIÓN", []),
    (2, "ARRENDAMIENTOS", []),
    (3, "Arrend. Locales",              ["5220101500"]),
    (3, "Maq. y Equipo",                ["5220150000"]),
    (2, "IMPUESTOS", []),
    (3, "ICA",                          ["5215050000"]),
    (3, "IVA Descontable",              ["5215700000"]),
    (2, "SS PP Y SERVICIOS", []),
    (3, "Energía Eléctrica",            ["5235300000"]),
    (3, "Agua y Alcantarillado",        ["5235250000"]),
    (3, "Teléfono Fijo",                ["5235350000"]),
    (3, "Celular",                      ["5235360000"]),
    (3, "Internet",                     ["5235380000"]),
    (3, "Gas Natural",                  ["5235390000"]),
    (2, "HONORARIOS", []),
    (3, "Honorarios Socios",            ["5210050000"]),
    (3, "Revisoría Fiscal",             ["5210100000"]),
    (3, "Otras Asesorías",              ["5210300000"]),
    (3, "Comisiones",                   ["5210010000"]),
    (2, "MANTENIMIENTO", []),
    (3, "Mtto Construcciones",          ["5245100000"]),
    (3, "Mtto Maquinaria",              ["5245150000"]),
    (3, "Mtto Cómputo",                 ["5245250000"]),
    (2, "DEPRECIACIONES", []),
    (3, "Maq. y Equipo",                ["5260200000"]),
    (0, "NO OPERACIONALES", []),
    (1, "INGRESOS NO OPER.", []),
    (3, "Recuper. Otros Gastos",        ["4250810000"]),
    (3, "Descuento Comercial",          ["4210401000"]),
    (3, "Descuento Financiero",         ["4210400500"]),
    (3, "Reintegro de Gastos",          ["4250500000"]),
    (3, "Intereses",                    ["4210050000"]),
    (1, "GASTOS NO OPER.", []),
    (3, "Multas y Sanciones",           ["5395200500"]),
    (3, "Pérd. Vta Activos",            ["5310150000"]),
    (0, "IMPUESTO RENTA", []),
    (3, "Renta y Complementarios",      ["5405050000"]),
    (3, "CREE",                         ["5405060000"]),
]

SALES_ACCOUNTS = [
    "4135220000","4135230000","4135240000","4135250000",
    "4135350500","4135351000","4135351500","4135352000","4135352500",
]
SALES_CATEGORIES = {
    "Fruver":    ["4135220000"],
    "Cárnicos":  ["4135230000","4135240000","4135250000"],
    "PGC":       ["4135350500","4135351000","4135351500","4135352000","4135352500"],
}


def _safe_float(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0


def _parse_account(raw):
    if raw is None: return None
    s = str(raw).strip().split(".")[0]
    return s if s.isdigit() else None


def _read_pyg_sheet(ws, month_cols, quarter_cols=None):
    q_def = {1:[1,2,3], 2:[4,5,6], 3:[7,8,9], 4:[10,11,12]}
    result = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) <= 3: continue
        cuenta = _parse_account(row[3])
        if not cuenta: continue
        acct = {f"M{m:02d}": _safe_float(row[c] if c < len(row) else None)
                for m, c in month_cols.items()}
        if quarter_cols:
            for q, c in quarter_cols.items():
                acct[f"Q{q}"] = _safe_float(row[c] if c < len(row) else None)
        else:
            for q, ms in q_def.items():
                acct[f"Q{q}"] = sum(acct.get(f"M{mm:02d}", 0) for mm in ms)
        result[cuenta] = acct
    return result


@lru_cache(maxsize=1)
def load_all():
    """Carga y cachea todos los datos de Excel al primer uso."""
    wb = openpyxl.load_workbook(FUENTE_FILE, keep_vba=False, data_only=True)
    ws = wb["FUENTE"]
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows) if r[0] == "ANIO")
    df = pd.DataFrame([dict(zip(rows[hi], r)) for r in rows[hi+1:] if r[0] is not None])
    for col in ("ANIO", "MES"):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["SALDO"]       = pd.to_numeric(df["SALDO"], errors="coerce").fillna(0)
    df["DEBE"]        = pd.to_numeric(df.get("DEBE", 0), errors="coerce").fillna(0)
    df["HABER"]       = pd.to_numeric(df.get("HABER", 0), errors="coerce").fillna(0)
    df["CUENTA"]      = df["CUENTA"].astype(str).str.strip()
    df["CENTROCOSTE"] = df["CENTROCOSTE"].astype(str).str.strip()
    df = df.dropna(subset=["ANIO", "MES"])

    ws2 = wb["centroscosto"]
    cr  = list(ws2.iter_rows(values_only=True))
    hcc = next(i for i, r in enumerate(cr) if r[0] == "CODALMACEN")
    cc_df = pd.DataFrame([r for r in cr[hcc+1:] if r[0] is not None],
                         columns=["CODALMACEN", "NOMBREALMACEN", "CENTROCOSTE"])
    cc_df["CENTROCOSTE"] = cc_df["CENTROCOSTE"].astype(str).str.strip()

    wb2 = openpyxl.load_workbook(PYG_FILE, keep_vba=False, data_only=True)
    pyg24 = pyg25 = {}
    for sn in wb2.sheetnames:
        su = sn.upper()
        if "TOTAL" in su and any(x in su for x in ("COMPAN", "COMPAÑ")):
            pyg24 = _read_pyg_sheet(wb2[sn], COMP_MONTH_COLS)
        elif "MERKACOL" in su and "CUENTA" in su:
            pyg25 = _read_pyg_sheet(wb2[sn], MERK_MONTH_COLS, MERK_QUARTER_COLS)

    lm_months_2026 = sorted(int(m) for m in df[df["ANIO"] == 2026]["MES"].unique())
    return df, cc_df, pyg24, pyg25, lm_months_2026
